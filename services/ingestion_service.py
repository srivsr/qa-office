"""
Ingestion service — multi-format test case parsing.
Supports: Excel (.xlsx/.xls), CSV (.csv), plain text (.txt/.text/.md).
Auto-detects format from extension; content-sniffs when extension is ambiguous.
Called by A1 Ingestion only. Never called directly by other agents.
"""

import csv
import io
import logging
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl

from schemas import TestCase

logger = logging.getLogger(__name__)

COLUMN_ALIASES: Dict[str, List[str]] = {
    "test_case_id": ["test case id", "tc id", "id", "test id", "#"],
    "module": ["module", "module name", "feature", "area"],
    "description": ["description", "test description", "test name", "summary"],
    "steps": ["test steps", "steps", "step description", "action"],
    "expected_result": [
        "expected result",
        "expected",
        "expected outcome",
        "expected behavior",
    ],
    "priority": ["priority"],
    "pre_conditions": [
        "pre-conditions",
        "preconditions",
        "pre conditions",
        "precondition",
    ],
}
IGNORE_COLUMNS = {"actual result", "actual", "comments", "remarks", "notes", "result"}
REQUIRED_FIELDS = {"test_case_id", "steps", "expected_result"}

_EXCEL_EXTS = {".xlsx", ".xls", ".xlsm"}
_CSV_EXTS = {".csv", ".tsv"}
_TEXT_EXTS = {".txt", ".text", ".md"}


class IngestionError(Exception):
    """Raised when a file cannot be parsed reliably — never guesses."""


# ── Public entry point ────────────────────────────────────────────────────────


def parse(path: str) -> List[TestCase]:
    """Auto-detect format and parse. Raises IngestionError on problems."""
    fmt = detect_format(path)
    logger.info("Detected format '%s' for %s", fmt, path)
    if fmt == "excel":
        return parse_excel(path)
    if fmt == "csv":
        return parse_csv(path)
    return parse_text(path)


def detect_format(path: str) -> str:
    """Return 'excel' | 'csv' | 'text'. Raises IngestionError for unsupported."""
    p = Path(path)
    if not p.exists():
        raise IngestionError(f"File not found: {path}")
    ext = p.suffix.lower()
    if ext in _EXCEL_EXTS:
        return "excel"
    if ext in _CSV_EXTS:
        return "csv"
    if ext in _TEXT_EXTS:
        return _sniff_text_or_csv(p)
    # Unknown extension — sniff content
    return _sniff_text_or_csv(p)


def _sniff_text_or_csv(p: Path) -> str:
    """Read first line; if it looks like delimited headers, treat as CSV."""
    try:
        first_line = p.read_text(encoding="utf-8", errors="replace").splitlines()[0]
    except IndexError:
        return "text"
    # Heuristic: CSV if first line has 3+ comma/tab/semicolon-separated tokens
    for sep in (",", "\t", ";"):
        if first_line.count(sep) >= 2:
            return "csv"
    return "text"


# ── Excel parser (unchanged from Phase 1) ────────────────────────────────────


def parse_excel(excel_path: str) -> List[TestCase]:
    """Parse BlueTree Excel → list[TestCase]. Raises IngestionError on problems."""
    path = Path(excel_path)
    if not path.exists():
        raise IngestionError(f"File not found: {excel_path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    all_tests: List[TestCase] = []
    for sheet_name in wb.sheetnames:
        tests = _parse_sheet(wb[sheet_name], sheet_name)
        all_tests.extend(tests)
        logger.info("Sheet '%s': %d test cases", sheet_name, len(tests))
    return all_tests


# ── CSV parser ────────────────────────────────────────────────────────────────


def parse_csv(csv_path: str) -> List[TestCase]:
    """
    Parse a CSV/TSV file into test cases.
    Auto-detects delimiter. Reuses COLUMN_ALIASES for header matching.
    Multi-line step cells must be newline-separated within the cell.
    """
    path = Path(csv_path)
    if not path.exists():
        raise IngestionError(f"File not found: {csv_path}")

    text = path.read_text(encoding="utf-8-sig", errors="replace")  # strip BOM
    dialect = _detect_csv_dialect(text)
    rows = list(csv.DictReader(io.StringIO(text), dialect=dialect))

    if not rows:
        raise IngestionError(f"CSV file is empty or has no data rows: {csv_path}")

    # Map header names → canonical field names using COLUMN_ALIASES
    raw_headers = {k.strip().lower(): k for k in rows[0].keys()}
    col_map = _match_csv_headers(raw_headers)

    missing = REQUIRED_FIELDS - set(col_map.keys())
    if missing:
        raise IngestionError(
            f"CSV missing required columns {missing}. "
            f"Found headers: {list(raw_headers.keys())}. Check aliases in DECISIONS.md ADR-014."
        )

    tests: List[TestCase] = []
    for i, row in enumerate(rows, start=2):  # row 2 = first data row

        def get(field: str) -> Optional[str]:
            raw_key = col_map.get(field)
            return (row.get(raw_key) or "").strip() or None if raw_key else None

        tc_id = get("test_case_id")
        steps_raw = get("steps")
        if not tc_id or not steps_raw:
            continue  # skip empty/incomplete rows

        pre = _split_steps(get("pre_conditions"))
        steps = pre + _split_steps(steps_raw)

        tests.append(
            TestCase(
                id=tc_id,
                module=get("module") or "General",
                description=get("description") or "",
                steps=steps,
                expected_result=get("expected_result") or "",
                priority=_normalise_priority(get("priority")),
                pre_conditions=pre,
            )
        )
        logger.debug("CSV row %d → TC %s (%d steps)", i, tc_id, len(steps))

    return tests


def _detect_csv_dialect(text: str) -> csv.Dialect:
    try:
        return csv.Sniffer().sniff(text[:2048], delimiters=",\t;")
    except csv.Error:
        return csv.excel  # fall back to comma


def _match_csv_headers(raw_headers: Dict[str, str]) -> Dict[str, str]:
    """Map canonical field name → original header string."""
    col_map: Dict[str, str] = {}
    for norm_header, orig_header in raw_headers.items():
        if norm_header in IGNORE_COLUMNS:
            continue
        for field, aliases in COLUMN_ALIASES.items():
            if norm_header in aliases and field not in col_map:
                col_map[field] = orig_header
                break
    return col_map


# ── Plain-text parser ─────────────────────────────────────────────────────────
#
# Accepts two text layouts:
#
# Layout A — structured blocks (QA team wrote proper test cases):
#   TC-001: Clock In Test
#   Steps:
#   1. Navigate to /attendance
#   2. Click Clock In
#   Expected: Attendance recorded
#   Priority: High
#
# Layout B — simple numbered list (quick requirement notes):
#   1. Open the app
#   2. Navigate to attendance tab
#   3. Click the Clock In button
#   Expected Result: Clock-in confirmed

_TC_HEADER = re.compile(r"^(TC[-\s]?\d+[\w-]*)\s*[:\-–]\s*(.*)", re.IGNORECASE)
_STEP_LINE = re.compile(r"^\s*(\d+)[.)]\s+(.+)")
_EXPECTED_LINE = re.compile(
    r"^\s*expected[^\w]*(result|outcome|behavior)?[:\s]+(.+)", re.IGNORECASE
)
_PRIORITY_LINE = re.compile(r"^\s*priority[:\s]+(.+)", re.IGNORECASE)
_PRECOND_LINE = re.compile(r"^\s*pre[- ]?condition[s]?[:\s]+(.+)", re.IGNORECASE)
_STEPS_SECTION = re.compile(r"^\s*steps[:\s]*$", re.IGNORECASE)


def parse_text(text_path: str) -> List[TestCase]:
    """
    Parse a plain-text file. Handles structured blocks (Layout A)
    and simple numbered lists (Layout B). Auto-detects layout.
    """
    path = Path(text_path)
    if not path.exists():
        raise IngestionError(f"File not found: {text_path}")

    lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
    # Detect layout: if any TC header line is found, use structured parser
    has_tc_headers = any(_TC_HEADER.match(ln) for ln in lines)
    if has_tc_headers:
        return _parse_text_structured(lines)
    return _parse_text_simple(lines, path.stem)


def _parse_text_structured(lines: List[str]) -> List[TestCase]:
    """Parse Layout A: explicit TC-ID blocks."""
    tests: List[TestCase] = []
    buf: Optional[dict] = None
    in_steps = False

    def flush() -> None:
        if buf and buf["steps"]:
            tests.append(_buf_to_testcase(buf))

    for line in lines:
        m_tc = _TC_HEADER.match(line)
        if m_tc:
            flush()
            buf = _empty_buf(m_tc.group(1), m_tc.group(2).strip())
            in_steps = False
            continue

        if buf is None:
            continue

        if _STEPS_SECTION.match(line):
            in_steps = True
            continue

        m_step = _STEP_LINE.match(line)
        if m_step:
            in_steps = True
            buf["steps"].append(m_step.group(2).strip())
            continue

        m_exp = _EXPECTED_LINE.match(line)
        if m_exp:
            buf["expected_result"] = m_exp.group(2).strip()
            in_steps = False
            continue

        m_pri = _PRIORITY_LINE.match(line)
        if m_pri:
            buf["priority"] = _normalise_priority(m_pri.group(1).strip())
            continue

        m_pre = _PRECOND_LINE.match(line)
        if m_pre:
            buf["pre_conditions"].append(m_pre.group(1).strip())
            continue

        # Continuation line inside steps block
        if in_steps and line.strip() and not line.startswith("TC"):
            buf["steps"].append(line.strip())

    flush()
    return tests


def _parse_text_simple(lines: List[str], file_stem: str) -> List[TestCase]:
    """
    Parse Layout B: plain numbered list with optional Expected line at end.
    Generates one TestCase per contiguous numbered-step block.
    """
    tests: List[TestCase] = []
    seq = 0
    buf: Optional[dict] = None

    def flush() -> None:
        nonlocal seq
        if buf and buf["steps"]:
            seq += 1
            buf["id"] = f"TC-{seq:03d}"
            tests.append(_buf_to_testcase(buf))

    for line in lines:
        m_step = _STEP_LINE.match(line)
        if m_step:
            if buf is None:
                buf = _empty_buf("", "")
            buf["steps"].append(m_step.group(2).strip())
            continue

        m_exp = _EXPECTED_LINE.match(line)
        if m_exp:
            if buf is None:
                buf = _empty_buf("", "")
            buf["expected_result"] = m_exp.group(2).strip()
            continue

        # Blank line between blocks → flush current
        if not line.strip() and buf and buf["steps"]:
            flush()
            buf = None

    flush()

    if not tests:
        raise IngestionError(
            "No test steps found in text file. "
            "Use numbered steps (1. Step text) or TC-001: headers."
        )
    return tests


def _empty_buf(tc_id: str, description: str) -> dict:
    return {
        "id": tc_id,
        "module": "General",
        "description": description,
        "steps": [],
        "expected_result": "",
        "priority": "Medium",
        "pre_conditions": [],
    }


def _buf_to_testcase(buf: dict) -> TestCase:
    steps = buf["pre_conditions"] + buf["steps"]
    return TestCase(
        id=buf["id"],
        module=buf["module"],
        description=buf["description"],
        steps=steps,
        expected_result=buf["expected_result"]
        or "Verify action completed successfully",
        priority=buf["priority"],
        pre_conditions=buf["pre_conditions"],
    )


# ── Shared helpers ────────────────────────────────────────────────────────────


def _split_steps(raw: Optional[str]) -> List[str]:
    return [s.strip() for s in raw.splitlines() if s.strip()] if raw else []


def _normalise_priority(raw: Optional[str]) -> str:
    if not raw:
        return "Medium"
    v = raw.strip().lower()
    if v in ("p1", "high", "critical"):
        return "High"
    if v in ("p3", "low"):
        return "Low"
    return "Medium"


# ── Excel internals (unchanged) ───────────────────────────────────────────────


def _parse_sheet(ws, sheet_name: str) -> List[TestCase]:
    merged = _resolve_merged_cells(ws)
    header_row, col_map = _find_headers(ws, merged)
    if header_row is None:
        logger.warning("Sheet '%s': no recognisable headers — skipping", sheet_name)
        return []
    missing = REQUIRED_FIELDS - set(col_map.keys())
    if missing:
        raise IngestionError(
            f"Sheet '{sheet_name}': missing required columns {missing}. "
            f"Recognised: {list(col_map.keys())}. Check aliases in REQUIREMENTS.md §5."
        )
    return _extract_test_cases(ws, merged, header_row, col_map, sheet_name)


def _resolve_merged_cells(ws) -> Dict[Tuple[int, int], object]:
    merge_map: Dict[Tuple[int, int], object] = {}
    for rng in ws.merged_cells.ranges:
        anchor = ws.cell(row=rng.min_row, column=rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                merge_map[(r, c)] = anchor
    result: Dict[Tuple[int, int], object] = {}
    for row in ws.iter_rows():
        for cell in row:
            key = (cell.row, cell.column)
            result[key] = merge_map.get(key, cell.value)
    return result


def _find_headers(ws, merged: Dict) -> Tuple[Optional[int], Dict[str, int]]:
    for row_num in range(1, min(11, ws.max_row + 1)):
        row_text = {
            col: str(merged.get((row_num, col)) or "").lower().strip()
            for col in range(1, ws.max_column + 1)
        }
        col_map = _match_columns(row_text)
        if len(col_map) >= 2:
            return row_num, col_map
    return None, {}


def _match_columns(row_text: Dict[int, str]) -> Dict[str, int]:
    col_map: Dict[str, int] = {}
    for col_idx, text in row_text.items():
        if not text or text in IGNORE_COLUMNS:
            continue
        for field, aliases in COLUMN_ALIASES.items():
            if text in aliases and field not in col_map:
                col_map[field] = col_idx
                break
    return col_map


def _extract_test_cases(ws, merged, header_row, col_map, sheet_name) -> List[TestCase]:
    tests: List[TestCase] = []
    current_id: Optional[str] = None
    buf: Optional[dict] = None

    def flush() -> None:
        if buf and buf["steps"]:
            tests.append(
                TestCase(
                    id=buf["id"],
                    module=buf["module"],
                    description=buf["description"],
                    steps=buf["pre_conditions"] + buf["steps"],
                    expected_result=buf["expected_result"],
                    priority=buf["priority"],
                    pre_conditions=buf["pre_conditions"],
                )
            )

    for row_num in range(header_row + 1, ws.max_row + 1):

        def cell(f: str) -> Optional[str]:
            c = col_map.get(f)
            v = merged.get((row_num, c)) if c else None
            return str(v).strip() if v is not None else None

        tc_id, steps_raw = cell("test_case_id"), cell("steps")
        if not tc_id and not steps_raw:
            continue
        if tc_id and tc_id != current_id:
            flush()
            current_id = tc_id
            buf = {
                "id": tc_id,
                "module": cell("module") or sheet_name,
                "description": cell("description") or "",
                "steps": [],
                "expected_result": cell("expected_result") or "",
                "priority": _normalise_priority(cell("priority")),
                "pre_conditions": _split_steps(cell("pre_conditions")),
            }
        if buf and steps_raw:
            buf["steps"].extend(_split_steps(steps_raw))
        if buf and not buf["expected_result"]:
            er = cell("expected_result")
            if er:
                buf["expected_result"] = er

    flush()
    return tests
