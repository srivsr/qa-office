"""
Excel Test Reporter Service

Generates Excel test reports with screenshot embedding.
Matches the template format: Test_case_template.xlsx
"""

import os
import base64
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Any
from dataclasses import dataclass
from enum import Enum
import io

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage
except ImportError:
    raise ImportError("openpyxl required. Install with: pip install openpyxl")

try:
    from PIL import Image
except ImportError:
    Image = None


class TestStatus(Enum):
    PENDING = "Pending"
    PASSED = "Pass"
    FAILED = "Fail"
    ERROR = "Error"
    BLOCKED = "Blocked"
    SKIPPED = "Skipped"


@dataclass
class TestCaseRow:
    """Test case row for Excel."""
    sl_no: int
    test_case_id: str
    requirement_id: str
    description: str
    expected_result: str
    status: TestStatus = TestStatus.PENDING
    technique: Optional[str] = None
    polarity: Optional[str] = None
    screenshot_path: Optional[str] = None
    screenshot_base64: Optional[str] = None
    comments: Optional[str] = None


class ExcelReporter:
    """Generates Excel test reports with screenshots."""

    # Styling
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    PENDING_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    ERROR_FILL = PatternFill(start_color="FF7043", end_color="FF7043", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    COLUMNS = {
        'A': ('Sl. NO', 8),
        'B': ('Test Case ID', 18),
        'C': ('Requirement ID', 18),
        'D': ('Description', 50),
        'E': ('Expected Result', 40),
        'F': ('Status', 12),
        'G': ('Screenshot', 35),
        'H': ('Comments', 40),
    }

    def __init__(self, output_dir: str = "test_results"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.screenshots_dir = self.output_dir / "screenshots"
        self.screenshots_dir.mkdir(parents=True, exist_ok=True)

    def generate_report(
        self,
        test_cases: List[TestCaseRow],
        app_name: str = "Application",
        embed_screenshots: bool = False
    ) -> bytes:
        """
        Generate Excel report and return as bytes.

        Args:
            test_cases: List of test case rows
            app_name: Application name for report
            embed_screenshots: If True, embed images; if False, add hyperlinks

        Returns:
            Excel file as bytes
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Cases"

        # Setup headers
        self._setup_headers(ws)

        # Add test cases
        for idx, tc in enumerate(test_cases, start=2):
            self._add_test_case_row(ws, idx, tc, embed_screenshots)

        # Add RTM sheet
        self._add_rtm_sheet(wb, test_cases)

        # Add Summary sheet
        self._add_summary_sheet(wb, test_cases, app_name)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def generate_report_file(
        self,
        test_cases: List[TestCaseRow],
        filename: str,
        app_name: str = "Application",
        embed_screenshots: bool = False
    ) -> str:
        """Generate report and save to file."""
        report_bytes = self.generate_report(test_cases, app_name, embed_screenshots)
        filepath = self.output_dir / filename
        with open(filepath, 'wb') as f:
            f.write(report_bytes)
        return str(filepath)

    def _setup_headers(self, ws):
        """Setup header row."""
        for col_letter, (header, width) in self.COLUMNS.items():
            cell = ws[f"{col_letter}1"]
            cell.value = header
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.THIN_BORDER
            ws.column_dimensions[col_letter].width = width

        ws.row_dimensions[1].height = 25
        ws.freeze_panes = 'A2'

    def _add_test_case_row(self, ws, row: int, tc: TestCaseRow, embed_screenshots: bool):
        """Add a test case row."""
        ws[f'A{row}'] = tc.sl_no
        ws[f'B{row}'] = tc.test_case_id
        ws[f'C{row}'] = tc.requirement_id
        ws[f'D{row}'] = tc.description
        ws[f'E{row}'] = tc.expected_result
        ws[f'F{row}'] = tc.status.value
        ws[f'H{row}'] = tc.comments or ""

        # Status coloring
        status_cell = ws[f'F{row}']
        if tc.status == TestStatus.PASSED:
            status_cell.fill = self.PASS_FILL
        elif tc.status == TestStatus.FAILED:
            status_cell.fill = self.FAIL_FILL
        elif tc.status == TestStatus.ERROR:
            status_cell.fill = self.ERROR_FILL
        elif tc.status == TestStatus.PENDING:
            status_cell.fill = self.PENDING_FILL

        # Default row height — set BEFORE screenshot so embed can override it
        ws.row_dimensions[row].height = 60

        # Screenshot handling
        if tc.screenshot_base64 or tc.screenshot_path:
            if embed_screenshots:
                self._embed_screenshot(ws, row, tc)
            else:
                path = tc.screenshot_path or "screenshot.png"
                ws[f'G{row}'] = path
                if tc.screenshot_path and os.path.exists(tc.screenshot_path):
                    ws[f'G{row}'].hyperlink = tc.screenshot_path
                    ws[f'G{row}'].font = Font(color="0563C1", underline="single")

        # Borders and alignment
        for col in 'ABCDEFGH':
            cell = ws[f'{col}{row}']
            cell.border = self.THIN_BORDER
            if col in ['A', 'B', 'C', 'F']:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    def _embed_screenshot(self, ws, row: int, tc: TestCaseRow):
        """Embed screenshot in cell."""
        if not Image:
            ws[f'G{row}'] = "PIL not installed"
            return

        try:
            if tc.screenshot_base64:
                img_data = base64.b64decode(tc.screenshot_base64)
                img = Image.open(io.BytesIO(img_data))
            elif tc.screenshot_path and os.path.exists(tc.screenshot_path):
                img = Image.open(tc.screenshot_path)
            else:
                ws[f'G{row}'] = "No screenshot"
                return

            # Resize to a clearly visible size (240×160 fits well in Excel)
            img.thumbnail((240, 160), Image.Resampling.LANCZOS)

            img_bytes = io.BytesIO()
            img.save(img_bytes, format='PNG')
            img_bytes.seek(0)

            xl_img = XLImage(img_bytes)
            xl_img.width = 240
            xl_img.height = 160
            ws.add_image(xl_img, f'G{row}')
            # Row height must be set AFTER add_image and large enough for the image
            # Excel row height is in points; 1pt ≈ 1.33px → 160px ÷ 1.33 ≈ 120pt
            ws.row_dimensions[row].height = 120

        except Exception as e:
            ws[f'G{row}'] = f"Error: {str(e)[:50]}"

    def _add_rtm_sheet(self, wb, test_cases: List[TestCaseRow]):
        """Add RTM sheet."""
        ws = wb.create_sheet("RTM")

        headers = ['Requirement ID', 'Test Case IDs', 'Total', 'Passed', 'Failed', 'Coverage %']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.THIN_BORDER

        # Group by requirement
        req_map: Dict[str, List[TestCaseRow]] = {}
        for tc in test_cases:
            if tc.requirement_id not in req_map:
                req_map[tc.requirement_id] = []
            req_map[tc.requirement_id].append(tc)

        row = 2
        for req_id, tcs in sorted(req_map.items()):
            tc_ids = ", ".join([tc.test_case_id for tc in tcs])
            total = len(tcs)
            passed = sum(1 for tc in tcs if tc.status == TestStatus.PASSED)
            failed = sum(1 for tc in tcs if tc.status == TestStatus.FAILED)
            coverage = (passed / total * 100) if total > 0 else 0

            ws.cell(row=row, column=1, value=req_id)
            ws.cell(row=row, column=2, value=tc_ids)
            ws.cell(row=row, column=3, value=total)
            ws.cell(row=row, column=4, value=passed)
            ws.cell(row=row, column=5, value=failed)
            ws.cell(row=row, column=6, value=f"{coverage:.1f}%")

            for col in range(1, 7):
                ws.cell(row=row, column=col).border = self.THIN_BORDER

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 12

    def _add_summary_sheet(self, wb, test_cases: List[TestCaseRow], app_name: str):
        """Add summary sheet."""
        ws = wb.create_sheet("Summary")

        total = len(test_cases)
        passed = sum(1 for tc in test_cases if tc.status == TestStatus.PASSED)
        failed = sum(1 for tc in test_cases if tc.status == TestStatus.FAILED)
        errored = sum(1 for tc in test_cases if tc.status == TestStatus.ERROR)
        pending = sum(1 for tc in test_cases if tc.status == TestStatus.PENDING)

        summary = [
            ("Test Execution Report", ""),
            ("Application", app_name),
            ("", ""),
            ("Total Test Cases", total),
            ("Passed", passed),
            ("Failed", failed),
            ("Error", errored),
            ("Pending", pending),
            ("", ""),
            ("Pass Rate", f"{(passed/total*100):.1f}%" if total > 0 else "N/A"),
            ("Execution Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]

        for row, (label, value) in enumerate(summary, 1):
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            if row == 1:
                ws.cell(row=row, column=1).font = Font(bold=True, size=14)
            elif label == "Passed":
                ws.cell(row=row, column=2).fill = self.PASS_FILL
            elif label == "Failed":
                ws.cell(row=row, column=2).fill = self.FAIL_FILL

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30

    def save_screenshot(self, screenshot_data: bytes, test_case_id: str) -> str:
        """Save screenshot and return path."""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{test_case_id}_{timestamp}.png"
        filepath = self.screenshots_dir / filename

        with open(filepath, 'wb') as f:
            f.write(screenshot_data)

        return str(filepath)


def generate_test_cases_excel(
    test_cases: List[Dict[str, Any]],
    app_name: str,
    session_id: str,
    output_dir: str = None
) -> str:
    """
    Generate Excel spreadsheet with test cases (before execution).
    Includes links to where test results will be stored.

    Args:
        test_cases: Generated test cases
        app_name: Application name
        session_id: Session ID for linking results
        output_dir: Output directory (defaults to app/data)

    Returns:
        Path to saved Excel file
    """
    if output_dir is None:
        output_dir = str(Path(__file__).parent.parent / "data")

    Path(output_dir).mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # Sheet 1: Test Cases
    ws = wb.active
    ws.title = "Test Cases"

    # Headers
    headers = ['Sl.No', 'Test Case ID', 'Requirement ID', 'Technique', 'Positive/Negative', 'Route', 'Title', 'Description', 'Steps', 'Expected Result', 'Priority', 'Status', 'Result Link']
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Column widths: A=Sl, B=TC ID, C=Req ID, D=Technique, E=Polarity, F=Route, G=Title, H=Desc, I=Steps, J=Expected, K=Priority, L=Status, M=Link
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 40
    ws.column_dimensions['H'].width = 50
    ws.column_dimensions['I'].width = 60
    ws.column_dimensions['J'].width = 40
    ws.column_dimensions['K'].width = 10
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 25

    # Technique display names
    TECHNIQUE_LABELS = {
        "ep_positive": "EP - Positive",
        "ep_negative": "EP - Negative",
        "equivalence_partition": "EP - Negative",
        "positive": "EP - Positive",
        "negative": "Negative",
        "bva": "Boundary Value (BVA)",
        "boundary": "Boundary Value (BVA)",
        "security": "Security",
        "usability": "Usability",
        "ctd": "Classification Tree (CTD)",
    }

    POLARITY_FILLS = {
        "Positive": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "Negative": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }

    def _get_polarity(tc):
        explicit = tc.get('polarity', '')
        if explicit in ('Positive', 'Negative'):
            return explicit
        technique = tc.get('technique', '')
        return 'Positive' if technique in ('ep_positive', 'positive', 'usability') else 'Negative'

    # Test case rows
    for idx, tc in enumerate(test_cases, 1):
        row = idx + 1
        tc_id = tc.get('id', f'TC-{idx:03d}')

        steps = tc.get('steps', [])
        steps_text = '\n'.join(f"{i+1}. {s}" for i, s in enumerate(steps) if s) if isinstance(steps, list) else str(steps)

        technique_raw = tc.get('technique', '')
        technique_label = TECHNIQUE_LABELS.get(technique_raw, technique_raw.replace('_', ' ').title())
        polarity = _get_polarity(tc)

        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=tc_id)
        ws.cell(row=row, column=3, value=tc.get('requirement_id', ''))
        ws.cell(row=row, column=4, value=technique_label)
        ws.cell(row=row, column=5, value=polarity)
        ws.cell(row=row, column=6, value=tc.get('route', '/'))
        ws.cell(row=row, column=7, value=tc.get('title', ''))
        ws.cell(row=row, column=8, value=tc.get('description', ''))
        ws.cell(row=row, column=9, value=steps_text)
        ws.cell(row=row, column=10, value=tc.get('expected_result', ''))
        ws.cell(row=row, column=11, value=tc.get('priority', 'Medium'))
        ws.cell(row=row, column=12, value='Pending')

        # Polarity colour coding
        ws.cell(row=row, column=5).fill = POLARITY_FILLS.get(polarity, POLARITY_FILLS['Negative'])

        # Result link
        screenshot_link = f"screenshots/{session_id}/{tc_id}.png"
        link_cell = ws.cell(row=row, column=13, value=screenshot_link)
        link_cell.font = Font(color="0563C1", underline="single")

        for col in range(1, 14):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

        ws.row_dimensions[row].height = 60

    # Sheet 2: Summary
    ws_summary = wb.create_sheet("Summary")
    summary_data = [
        ("Test Execution Plan", ""),
        ("Application", app_name),
        ("Session ID", session_id),
        ("Total Test Cases", len(test_cases)),
        ("Generated Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("", ""),
        ("Route Distribution", ""),
    ]

    # Count routes
    route_counts = {}
    for tc in test_cases:
        route = tc.get('route', '/')
        route_counts[route] = route_counts.get(route, 0) + 1

    for route, count in sorted(route_counts.items(), key=lambda x: -x[1]):
        summary_data.append((route, count))

    for row, (label, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row, column=1, value=label)
        ws_summary.cell(row=row, column=2, value=value)
        if row == 1:
            ws_summary.cell(row=row, column=1).font = Font(bold=True, size=14)

    ws_summary.column_dimensions['A'].width = 30
    ws_summary.column_dimensions['B'].width = 40

    # Save file
    safe_app_name = "".join(c if c.isalnum() or c in "- _" else "_" for c in app_name)
    filename = f"test_cases_{safe_app_name}_{session_id}.xlsx"
    filepath = Path(output_dir) / filename

    wb.save(str(filepath))
    print(f"[EXCEL] Test cases saved to: {filepath}")

    return str(filepath)


def convert_test_results_to_rows(
    test_results: List[Dict[str, Any]],
    requirements: List[Dict[str, Any]]
) -> List[TestCaseRow]:
    """
    Convert test execution results to TestCaseRow objects.

    Args:
        test_results: Results from test execution
        requirements: Original requirements with IDs

    Returns:
        List of TestCaseRow for Excel generation
    """
    rows = []
    req_map = {r.get('req_id', ''): r for r in requirements}

    for idx, result in enumerate(test_results, 1):
        req_id = result.get('requirement_id', result.get('req_id', f'REQ-{idx:03d}'))

        # Determine status
        status_str = result.get('status', 'pending').lower()
        if status_str in ['pass', 'passed', 'success']:
            status = TestStatus.PASSED
        elif status_str in ['fail', 'failed', 'error']:
            status = TestStatus.FAILED
        elif status_str in ['skip', 'skipped']:
            status = TestStatus.SKIPPED
        elif status_str in ['block', 'blocked']:
            status = TestStatus.BLOCKED
        else:
            status = TestStatus.PENDING

        technique_raw = result.get('technique', '')
        polarity = result.get('polarity') or ('Positive' if technique_raw in ('ep_positive', 'positive', 'usability') else 'Negative')

        rows.append(TestCaseRow(
            sl_no=idx,
            test_case_id=result.get('test_id', result.get('id', f'TC-{idx:03d}')),
            requirement_id=req_id,
            description=result.get('description', result.get('title', '')),
            expected_result=result.get('expected_result', result.get('expected', '')),
            status=status,
            technique=technique_raw,
            polarity=polarity,
            screenshot_path=result.get('screenshot_path'),
            screenshot_base64=result.get('screenshot_base64'),
            comments=result.get('error_message') or result.get('comments', ''),
        ))

    return rows
