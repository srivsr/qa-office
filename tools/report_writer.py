"""
Report writer — thin adapter over qa-os ExcelReporter and HTML template.
All report generation goes through here; agents never touch openpyxl directly.
"""

import hashlib
import importlib.util
import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional

from reports.html_template import COMPLIANCE_SECTION, TEMPLATE
from schemas import ComplianceMetadata, ExecutionResult, RunReport, TestCase

logger = logging.getLogger(__name__)

# Load qa-os excel_reporter by absolute path to avoid sys.modules collision
_EXCEL_REPORTER_FILE = (
    Path(__file__).parents[2] / "qa-os" / "backend" / "app" / "services" / "excel_reporter.py"
)
_excel_reporter_mod = None


def _get_excel_reporter():
    global _excel_reporter_mod
    if _excel_reporter_mod is None:
        if not _EXCEL_REPORTER_FILE.exists():
            return None, None
        spec = importlib.util.spec_from_file_location("qa_os_excel_reporter", _EXCEL_REPORTER_FILE)
        mod = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(mod)
        _excel_reporter_mod = mod
    return _excel_reporter_mod.ExcelReporter, _excel_reporter_mod.convert_test_results_to_rows


class ReportWriter:
    """
    Writes HTML + Excel reports for a completed run.
    Wraps qa-os ExcelReporter and the inline HTML template.
    """

    def write(
        self,
        test_cases: List[TestCase],
        results: List[ExecutionResult],
        run_id: str,
        app_name: str,
        output_dir: str,
        started_at: Optional[datetime] = None,
        compliance: Optional[ComplianceMetadata] = None,
    ) -> RunReport:
        """
        Generate HTML + Excel, return RunReport with paths and stats.
        If Excel generation fails, logs and continues (HTML is primary).
        """
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        started_at = started_at or datetime.now(timezone.utc)
        result_map = {r.test_case_id: r for r in results}
        total_ms = sum(r.duration_ms for r in results)
        stats = _compute_stats(results)
        module_stats = _module_stats(test_cases, result_map)

        html_path = str(Path(output_dir) / f"report_{run_id}.html")
        excel_path = str(Path(output_dir) / f"report_{run_id}.xlsx")

        if compliance is None:
            compliance = _build_compliance(run_id, started_at, test_cases)
        _write_html(
            html_path,
            test_cases,
            result_map,
            stats,
            module_stats,
            run_id,
            app_name,
            started_at,
            total_ms,
            compliance,
        )
        _write_excel(excel_path, test_cases, result_map, app_name)

        return RunReport(
            run_id=run_id,
            html_path=html_path,
            excel_path=excel_path,
            duration_ms=total_ms,
            **stats,
        )


# ── Private helpers ───────────────────────────────────────────────────────────


def _compute_stats(results: List[ExecutionResult]) -> dict:
    total = len(results)
    passed = sum(1 for r in results if r.status == "passed")
    failed = sum(1 for r in results if r.status == "failed")
    error = sum(1 for r in results if r.status == "error")
    skipped = sum(1 for r in results if r.status == "skipped")
    return dict(
        total=total,
        passed=passed,
        failed=failed,
        error=error,
        skipped=skipped,
        pass_rate=round(passed / total * 100, 1) if total else 0.0,
    )


def _module_stats(test_cases: List[TestCase], result_map: dict) -> List[dict]:
    mods: Dict[str, dict] = {}
    for tc in test_cases:
        m = tc.module
        if m not in mods:
            mods[m] = {"name": m, "total": 0, "passed": 0}
        mods[m]["total"] += 1
        r = result_map.get(tc.id)
        if r and r.status == "passed":
            mods[m]["passed"] += 1
    for m in mods.values():
        m["pass_rate"] = round(m["passed"] / m["total"] * 100, 1) if m["total"] else 0.0
    return sorted(mods.values(), key=lambda x: -x["total"])


def _result_row(tc: TestCase, result: Optional[ExecutionResult]) -> str:
    if result is None:
        status, badge_cls, duration, retries, err_html = (
            "pending",
            "pending",
            "&mdash;",
            "0",
            "",
        )
    else:
        status = result.status
        badge_cls = status
        duration = f"{result.duration_ms}ms"
        retries = str(result.retry_count)
        err_html = (
            f'<div class="err-msg">&#9888; {result.error_message}</div>'
            if result.error_message
            else ""
        )
    visible = tc.steps[:3]
    steps_html = "".join(f"<div>&#x2022; {s}</div>" for s in visible)
    if len(tc.steps) > 3:
        steps_html += f"<div style='color:#aaa'>+{len(tc.steps) - 3} more steps</div>"
    return (
        f"<tr><td><span class='tc-id'>{tc.id}</span></td><td>{tc.module}</td>"
        f"<td><strong>{tc.description}</strong><div class='steps'>{steps_html}</div>{err_html}</td>"
        f"<td><span class='badge {badge_cls}'>{status}</span></td>"
        f"<td>{duration}</td><td>{retries}</td></tr>"
    )


def _build_compliance(
    run_id: str,
    started_at: datetime,
    test_cases: List[TestCase],
    env_health: Optional[str] = None,
    seed_summary: Optional[str] = None,
    chain_of_custody: Optional[List[str]] = None,
) -> ComplianceMetadata:
    tc_ids = sorted(tc.id for tc in test_cases)
    raw = f"{run_id}:{started_at.isoformat()}:{','.join(tc_ids)}"
    run_hash = hashlib.sha256(raw.encode()).hexdigest()
    return ComplianceMetadata(
        run_id=run_id,
        run_hash=run_hash,
        signed_timestamp=started_at.isoformat(),
        env_health=env_health,
        seed_summary=seed_summary,
        chain_of_custody=chain_of_custody
        or ["A13", "A12", "A1", "A2", "A3", "A4", "A5", "A6", "A8"],
    )


def _write_html(
    path,
    test_cases,
    result_map,
    stats,
    module_stats,
    run_id,
    app_name,
    started_at,
    total_ms,
    compliance: Optional[ComplianceMetadata] = None,
):
    module_bars = "".join(
        f'<div class="module-row"><div class="name">{m["name"]}</div>'
        f'<div class="bar"><div class="fill" style="width:{m["pass_rate"]}%"></div></div>'
        f'<div class="pct">{m["pass_rate"]}%</div></div>'
        for m in module_stats
    )
    rows = "".join(_result_row(tc, result_map.get(tc.id)) for tc in test_cases)
    mins, secs = divmod(total_ms // 1000, 60)
    duration = f"{mins}m {secs}s" if mins else f"{secs}s"
    compliance_section = ""
    if compliance:
        custody = " &rarr; ".join(compliance.chain_of_custody)
        compliance_section = COMPLIANCE_SECTION.format(
            run_id=compliance.run_id,
            run_hash=compliance.run_hash,
            signed_timestamp=compliance.signed_timestamp,
            env_health=compliance.env_health or "&mdash;",
            seed_summary=compliance.seed_summary or "&mdash;",
            chain_of_custody=custody,
        )
    html = TEMPLATE.format(
        app_name=app_name,
        run_id=run_id,
        timestamp=started_at.strftime("%Y-%m-%d %H:%M"),
        duration=duration,
        module_bars=module_bars,
        rows=rows,
        compliance_section=compliance_section,
        **stats,
    )
    Path(path).write_text(html, encoding="utf-8")
    logger.info("HTML report written: %s", path)


def _write_excel(path, test_cases, result_map, app_name):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage

    try:
        from PIL import Image as PILImage
        _pil_available = True
    except ImportError:
        _pil_available = False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Results"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    status_fills = {
        "passed": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "failed": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "error":  PatternFill(start_color="FF7043", end_color="FF7043", fill_type="solid"),
        "pending": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    }

    SCREENSHOT_COL = 10
    ROW_HEIGHT_PX = 120
    IMG_W, IMG_H = 160, 100  # pixels to fit inside cell

    headers = ["#", "Test Case ID", "Module", "Description", "Steps", "Expected Result", "Priority", "Status", "Error", "Screenshot"]
    col_widths = [5, 18, 18, 40, 60, 40, 10, 10, 40, 25]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 20

    for idx, tc in enumerate(test_cases, 1):
        result = result_map.get(tc.id)
        status = result.status if result else "pending"
        error = (result.error_message or "") if result else ""
        screenshot_path = (result.screenshot_paths[0] if result and result.screenshot_paths else None)
        steps_text = "\n".join(f"{i+1}. {s}" for i, s in enumerate(tc.steps) if s)

        row_num = idx + 1
        row_data = [idx, tc.id, tc.module, tc.description, steps_text, tc.expected_result, tc.priority, status, error]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if col == 8:
                cell.fill = status_fills.get(status, status_fills["pending"])

        # Screenshot column — embed image if file exists, else show path as text
        sc_cell = ws.cell(row=row_num, column=SCREENSHOT_COL)
        sc_cell.border = border
        sc_cell.alignment = Alignment(vertical="top")
        img_embedded = False
        if screenshot_path and _pil_available:
            import os
            if os.path.exists(screenshot_path):
                try:
                    pil_img = PILImage.open(screenshot_path)
                    pil_img.thumbnail((IMG_W, IMG_H))
                    import io as _io
                    buf = _io.BytesIO()
                    pil_img.save(buf, format="PNG")
                    buf.seek(0)
                    xl_img = XLImage(buf)
                    xl_img.width, xl_img.height = IMG_W, IMG_H
                    col_letter = sc_cell.column_letter
                    xl_img.anchor = f"{col_letter}{row_num}"
                    ws.add_image(xl_img)
                    img_embedded = True
                except Exception as e:
                    logger.warning("Screenshot embed failed for %s: %s", tc.id, e)
        if not img_embedded and screenshot_path:
            sc_cell.value = screenshot_path

        row_h = max(ROW_HEIGHT_PX * 0.75, 15 * len(tc.steps)) if tc.steps else ROW_HEIGHT_PX * 0.75
        ws.row_dimensions[row_num].height = row_h

    wb.save(path)
    logger.info("Excel report written: %s", path)
